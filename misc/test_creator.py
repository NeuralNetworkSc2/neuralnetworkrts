import openpyxl as opxl
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwt
import xlrd
import os, fnmatch
import random

def find(extension, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, extension):
                temp = os.path.join(root, name).split('\\')
                result.append(temp[-1])
    return result

print('Set the path to the map folder: ')
path = input()
all_maps = find('*.SC2Map', path)
map_count = len([map for map in all_maps if map != 0])
if map_count == 0:
    print('In this directory no maps, try again')
else:
    print(map_count)
print(all_maps)

workbook = opxl.load_workbook('Book1.xlsx')
current_sheet = workbook['Sheet1']
print(current_sheet.max_row)
print(current_sheet.max_column)
for i in range(1, current_sheet.max_row + 1 ):
    for j in range(1, current_sheet.max_column + 1):
        print(current_sheet.cell(row=i, column=j).value)

num_of_tests = int(input("Enter number of tests for bot: "))
difficulties = []
difficulties.append("VeryEasy")
difficulties.append("Easy")
difficulties.append("Medium")
difficulties.append("MediumHard")
difficulties.append("Hard")
difficulties.append("Harder")
difficulties.append("VeryHard")
races = ["Protoss", "Terran", "Zerg"]

book = xlwt.Workbook(encoding="utf-8")
main_sheet = book.add_sheet("Sheet1")
for i in range(num_of_tests):
    map = random.choice(all_maps)
    difficulty = random.choice(difficulties)
    race = random.choice(races)
    main_sheet.write(i, 0, map)
    main_sheet.write(i, 1, "VeryEasy")
    main_sheet.write(i, 2, race)
book.save("tests_set.xls")

test_workbook = xlrd.open_workbook_xls("tests_set.xls")
current_sheet_test = test_workbook["Sheet1"]
c = current_sheet_test.cell(0, 0)
print(c)
